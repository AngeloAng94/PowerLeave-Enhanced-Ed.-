#!/usr/bin/env python3
"""
Script di importazione dati da Excel PianoFerie2025.xlsx
Struttura identificata:
- Riga 8: Mesi (date come 2021-04-01, 2021-05-01...)
- Riga 9: Giorni (date complete: 2021-04-01, 2021-04-02...) + "Team Member" in colonna B
- Colonna B (da riga 10): Nomi dipendenti
- Celle dati: X (8H), P2 (2H), P4 (4H)
"""

import openpyxl
import mysql.connector
from datetime import datetime, date
import os
from dotenv import load_dotenv

# Carica variabili ambiente
load_dotenv()

# Configurazione database
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'gateway01.us-west-2.prod.aws.tidbcloud.com'),
    'port': int(os.getenv('DB_PORT', 4000)),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME'),
    'ssl_ca': '/etc/ssl/certs/ca-certificates.crt',
    'ssl_verify_cert': True,
    'ssl_verify_identity': True
}

# Parsing DATABASE_URL se disponibile
if os.getenv('DATABASE_URL'):
    import re
    url = os.getenv('DATABASE_URL')
    match = re.match(r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/([^\?]+)', url)
    if match:
        DB_CONFIG['user'] = match.group(1)
        DB_CONFIG['password'] = match.group(2)
        DB_CONFIG['host'] = match.group(3)
        DB_CONFIG['port'] = int(match.group(4))
        DB_CONFIG['database'] = match.group(5)

# Mappatura tipi assenza
LEAVE_TYPE_MAP = {
    'X': ('FERIE/ROL 8 H', 8),
    'P2': ('PERMESSO 2 ORE', 2),
    'P4': ('PERMESSO 4 ORE', 4)
}

def parse_excel_date(cell_value):
    """Converte valore cella Excel in oggetto date"""
    if isinstance(cell_value, datetime):
        return cell_value.date()
    elif isinstance(cell_value, date):
        return cell_value
    elif isinstance(cell_value, str):
        try:
            return datetime.strptime(cell_value, '%Y-%m-%d').date()
        except:
            return None
    return None

def clean_db(conn):
    """Pulisce database prima dell'import"""
    cursor = conn.cursor()
    print("🧹 Pulizia database...")
    
    cursor.execute("DELETE FROM leave_requests")
    cursor.execute("DELETE FROM leave_balances")
    cursor.execute("DELETE FROM users WHERE role != 'admin'")
    cursor.execute("DELETE FROM company_closures")
    
    conn.commit()
    cursor.close()
    print("✓ Database pulito")

def import_excel(excel_path, conn):
    """Importa dati da Excel al database"""
    
    print(f"📊 Apertura file Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"✓ Foglio caricato: {ws.max_row} righe, {ws.max_column} colonne\n")
    
    # Passo 1: Leggi intestazione giorni (riga 9, da colonna 3)
    print("📅 Lettura intestazione date...")
    dates = []
    col_idx = 3  # Colonna C
    while col_idx <= ws.max_column:
        cell = ws.cell(row=9, column=col_idx)
        day_date = parse_excel_date(cell.value)
        if day_date:
            dates.append((col_idx, day_date))
        col_idx += 1
    
    print(f"✓ Trovate {len(dates)} date (da {dates[0][1] if dates else 'N/A'} a {dates[-1][1] if dates else 'N/A'})")
    
    # Passo 2: Leggi nomi dipendenti (colonna B, da riga 10)
    print("\n👥 Lettura nomi dipendenti...")
    employees = []
    row_idx = 10
    while row_idx <= ws.max_row:
        cell = ws.cell(row=row_idx, column=2)  # Colonna B
        if cell.value and str(cell.value).strip():
            name = str(cell.value).strip()
            if name not in ['Team Member', 'X', 'P2', 'P4']:
                employees.append((row_idx, name))
        row_idx += 1
    
    print(f"✓ Trovati {len(employees)} dipendenti")
    for _, name in employees[:5]:
        print(f"   - {name}")
    if len(employees) > 5:
        print(f"   ... e altri {len(employees) - 5}")
    
    # Passo 3: Crea utenti nel database
    print("\n💾 Creazione utenti nel database...")
    cursor = conn.cursor(buffered=True)
    user_map = {}  # nome -> user_id
    
    for _, emp_name in employees:
        # Genera openId univoco
        open_id = f"excel-import-{emp_name.lower().replace(' ', '-')}"
        
        cursor.execute("""
            INSERT INTO users (openId, name, email, role, createdAt, updatedAt, lastSignedIn)
            VALUES (%s, %s, %s, 'user', NOW(), NOW(), NOW())
            ON DUPLICATE KEY UPDATE name = VALUES(name)
        """, (open_id, emp_name, f"{emp_name.lower().replace(' ', '.')}@company.com"))
        
        # Ottieni user_id
        cursor.execute("SELECT id FROM users WHERE openId = %s", (open_id,))
        user_id = cursor.fetchone()[0]
        user_map[emp_name] = user_id
    
    conn.commit()
    print(f"✓ Creati {len(user_map)} utenti")
    
    # Passo 4: Crea tipi di assenza
    print("\n📋 Creazione tipi di assenza...")
    leave_type_ids = {}
    
    for code, (name, hours) in LEAVE_TYPE_MAP.items():
        cursor.execute("""
            INSERT INTO leave_types (name, description, requires_approval, createdAt)
            VALUES (%s, %s, 1, NOW())
            ON DUPLICATE KEY UPDATE name = VALUES(name)
        """, (name, f"{code} - {hours} ore"))
        
        cursor.execute("SELECT id FROM leave_types WHERE name = %s", (name,))
        result = cursor.fetchone()
        if result:
            leave_type_ids[code] = result[0]
    
    conn.commit()
    print(f"✓ Creati {len(leave_type_ids)} tipi di assenza")
    
    # Passo 5: Importa richieste ferie
    print("\n🏖️  Importazione richieste ferie...")
    requests_count = 0
    closures_dates = {}  # data -> count (per rilevare chiusure aziendali)
    
    for emp_row, emp_name in employees:
        user_id = user_map[emp_name]
        
        for col_idx, day_date in dates:
            cell = ws.cell(row=emp_row, column=col_idx)
            cell_value = str(cell.value).strip().upper() if cell.value else None
            
            if cell_value in LEAVE_TYPE_MAP:
                leave_code = cell_value
                leave_type_id = leave_type_ids[leave_code]
                _, hours = LEAVE_TYPE_MAP[leave_code]
                
                # Inserisci richiesta (già approvata perché storica)
                cursor.execute("""
                    INSERT INTO leave_requests 
                    (userId, leaveTypeId, startDate, endDate, days, hours, status, notes, createdAt, updatedAt)
                    VALUES (%s, %s, %s, %s, 1, %s, 'approved', 'Importato da Excel', NOW(), NOW())
                """, (user_id, leave_type_id, str(day_date), str(day_date), hours))
                
                requests_count += 1
                
                # Conta per rilevare chiusure aziendali
                if day_date not in closures_dates:
                    closures_dates[day_date] = 0
                closures_dates[day_date] += 1
    
    conn.commit()
    print(f"✓ Importate {requests_count} richieste di ferie")
    
    # Passo 6: Rileva e salva chiusure aziendali (giorni con tutti/quasi tutti in ferie)
    print("\n🏢 Rilevamento chiusure aziendali...")
    total_employees = len(employees)
    threshold = max(1, int(total_employees * 0.8))  # 80% dei dipendenti
    
    closures_count = 0
    for day_date, count in closures_dates.items():
        if count >= threshold:
            cursor.execute("""
                INSERT INTO company_closures (date, reason, type, createdAt)
                VALUES (%s, %s, 'holiday', NOW())
                ON DUPLICATE KEY UPDATE reason = VALUES(reason)
            """, (str(day_date), f"Chiusura aziendale ({count}/{total_employees} dipendenti)"))
            closures_count += 1
    
    conn.commit()
    print(f"✓ Rilevate {closures_count} chiusure aziendali (soglia: {threshold}/{total_employees} dipendenti)")
    
    # Passo 7: Calcola saldi ferie
    print("\n💰 Calcolo saldi ferie...")
    
    # Ottieni ID tipo ferie principale (FERIE/ROL 8H)
    ferie_type_id = leave_type_ids.get('X')
    
    for emp_name, user_id in user_map.items():
        # Calcola giorni utilizzati (conta richieste approvate)
        cursor.execute("""
            SELECT COALESCE(COUNT(*), 0)
            FROM leave_requests
            WHERE userId = %s AND status = 'approved' AND leaveTypeId = %s
        """, (user_id, ferie_type_id))
        used_days = int(cursor.fetchone()[0])
        
        # Saldo iniziale standard (26 giorni)
        total_days = 26
        
        cursor.execute("""
            INSERT INTO leave_balances (user_id, leave_type_id, total_days, used_days, year, createdAt, updatedAt)
            VALUES (%s, %s, %s, %s, 2025, NOW(), NOW())
            ON DUPLICATE KEY UPDATE 
                used_days = VALUES(used_days),
                updatedAt = NOW()
        """, (user_id, ferie_type_id, total_days, used_days))
    
    conn.commit()
    print(f"✓ Calcolati saldi per {len(user_map)} dipendenti")
    
    cursor.close()
    
    return {
        'employees': len(employees),
        'dates': len(dates),
        'requests': requests_count,
        'closures': closures_count
    }

def main():
    excel_path = '/home/ubuntu/upload/PianoFerie2025.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ File non trovato: {excel_path}")
        return
    
    print("=" * 60)
    print("  IMPORTAZIONE DATI EXCEL -> POWER LEAVE")
    print("=" * 60)
    print()
    
    try:
        # Connessione database
        print("🔌 Connessione al database...")
        conn = mysql.connector.connect(**DB_CONFIG)
        print("✓ Connesso\n")
        
        # Pulizia database
        clean_db(conn)
        print()
        
        # Importazione
        stats = import_excel(excel_path, conn)
        
        # Chiusura
        conn.close()
        
        # Riepilogo
        print("\n" + "=" * 60)
        print("🎉 IMPORTAZIONE COMPLETATA CON SUCCESSO!")
        print("=" * 60)
        print(f"   📊 Dipendenti importati:     {stats['employees']}")
        print(f"   📅 Date elaborate:           {stats['dates']}")
        print(f"   🏖️  Richieste ferie:          {stats['requests']}")
        print(f"   🏢 Chiusure aziendali:       {stats['closures']}")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ ERRORE: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
