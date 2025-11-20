"""
Import corretto da Excel con struttura reale:
- Riga 1: Mesi (gen-25, feb-25, ecc.)
- Riga 2: Giorni (1, 2, 3, ..., 31)
- Colonna A (da riga 3): Nomi dipendenti
- Celle: X, P2, P4
- Celle gialle: Chiusure aziendali
"""
import openpyxl
import mysql.connector
from datetime import datetime
from collections import defaultdict
import os
from dotenv import load_dotenv

# Carica .env.import
load_dotenv('/home/ubuntu/dashboard-ferie-team/.env.import')

# Connessione database
conn = mysql.connector.connect(
    host=os.getenv('DB_HOST'),
    port=int(os.getenv('DB_PORT', 4000)),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    database=os.getenv('DB_NAME'),
    ssl_ca=os.getenv('DB_SSL_CA') if os.getenv('DB_SSL_CA') else None
)
cursor = conn.cursor()

# Apri Excel
wb = openpyxl.load_workbook('/home/ubuntu/dashboard-ferie-team/PianoFerie2025.xlsx')
ws = wb.active

print("📊 Analisi struttura Excel...")

# Riga 1: Mesi, Riga 2: Giorni
# Costruisci mappa colonna → data
dates_map = {}
for col_idx in range(2, ws.max_column + 1):  # Dalla colonna B in poi
    month_cell = ws.cell(1, col_idx).value  # gen-25, feb-25, ecc.
    day_cell = ws.cell(2, col_idx).value    # 1, 2, 3, ..., 31
    
    if month_cell and day_cell:
        # Parsing mese-anno (es. "gen-25" → 2025-01)
        month_str = str(month_cell).strip().lower()
        month_map = {
            'gen': '01', 'feb': '02', 'mar': '03', 'apr': '04',
            'mag': '05', 'giu': '06', 'lug': '07', 'ago': '08',
            'set': '09', 'ott': '10', 'nov': '11', 'dic': '12'
        }
        
        for it_month, num_month in month_map.items():
            if it_month in month_str:
                # Estrai anno (es. "gen-25" → "25")
                year_part = month_str.split('-')[-1] if '-' in month_str else '25'
                year = f"20{year_part}" if len(year_part) == 2 else year_part
                
                # Costruisci data
                day = str(day_cell).strip()
                if day.isdigit():
                    date_str = f"{year}-{num_month}-{int(day):02d}"
                    dates_map[col_idx] = date_str
                break

print(f"✓ Trovate {len(dates_map)} date")
if dates_map:
    first_date = min(dates_map.values())
    last_date = max(dates_map.values())
    print(f"  Range: {first_date} → {last_date}")

# Leggi tutti i dipendenti (dalla riga 3 in poi, colonna A)
employees = {}
row_idx = 3
while True:
    name_cell = ws.cell(row_idx, 1).value  # Colonna A
    if not name_cell or str(name_cell).strip() == '':
        break
    
    name = str(name_cell).strip()
    
    # Crea utente
    cursor.execute(
        "INSERT INTO users (openId, name, email, role) VALUES (%s, %s, %s, %s) ON DUPLICATE KEY UPDATE name=VALUES(name)",
        (f"emp_{name.lower().replace(' ', '_')}", name, f"{name.lower().replace(' ', '.')}@company.com", "user")
    )
    conn.commit()
    
    # Get user ID
    cursor.execute("SELECT id FROM users WHERE openId = %s", (f"emp_{name.lower().replace(' ', '_')}",))
    user_id = cursor.fetchone()[0]
    employees[row_idx] = {'id': user_id, 'name': name}
    
    row_idx += 1

print(f"✓ Creati {len(employees)} dipendenti")

# Get leave types
cursor.execute("SELECT id, name FROM leave_types")
leave_types = {name: id for id, name in cursor.fetchall()}
print(f"✓ Trovati {len(leave_types)} tipi di assenza")

# Raccogli ferie per dipendente
employee_leaves = defaultdict(lambda: defaultdict(list))
company_closures = set()

for row_idx, emp_data in employees.items():
    for col_idx, date_str in dates_map.items():
        cell = ws.cell(row_idx, col_idx)
        cell_value = cell.value
        
        # Check se è chiusura aziendale (cella gialla)
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color.rgb
            if color and 'FFFF' in str(color):  # Giallo
                company_closures.add(date_str)
        
        if cell_value in ['X', 'x', 'P2', 'P4']:
            # Determina tipo e ore
            val = str(cell_value).upper()
            if val == 'X':
                leave_type = 'FERIE/ROL 8H'
                hours = 8
            elif val == 'P2':
                leave_type = 'PERMESSO 2 ORE'
                hours = 2
            else:  # P4
                leave_type = 'PERMESSO 4 ORE'
                hours = 4
            
            employee_leaves[emp_data['id']][leave_type].append((date_str, hours))

print(f"✓ Trovate {len(company_closures)} chiusure aziendali")

# Inserisci richieste raggruppate
total_requests = 0
for user_id, types_data in employee_leaves.items():
    for leave_type_name, dates_hours in types_data.items():
        leave_type_id = leave_types.get(leave_type_name)
        if not leave_type_id:
            continue
        
        # Raggruppa date consecutive con stesse ore
        dates_hours.sort()
        groups = []
        current_group = [dates_hours[0]]
        
        for i in range(1, len(dates_hours)):
            prev_date = datetime.strptime(current_group[-1][0], '%Y-%m-%d')
            curr_date = datetime.strptime(dates_hours[i][0], '%Y-%m-%d')
            diff = (curr_date - prev_date).days
            
            if diff == 1 and dates_hours[i][1] == current_group[-1][1]:
                current_group.append(dates_hours[i])
            else:
                groups.append(current_group)
                current_group = [dates_hours[i]]
        groups.append(current_group)
        
        # Inserisci ogni gruppo come richiesta
        for group in groups:
            start_date = group[0][0]
            end_date = group[-1][0]
            hours = group[0][1]
            days = len(group)
            
            cursor.execute("""
                INSERT INTO leave_requests 
                (userId, leaveTypeId, startDate, endDate, days, hours, status, notes)
                VALUES (%s, %s, %s, %s, %s, %s, 'approved', 'Importato da Excel')
            """, (user_id, leave_type_id, start_date, end_date, days, hours))
            total_requests += 1

conn.commit()
print(f"✓ Inserite {total_requests} richieste di ferie")

# Inserisci chiusure aziendali
for closure_date in company_closures:
    cursor.execute("""
        INSERT INTO company_closures (date, reason, type)
        VALUES (%s, 'Chiusura aziendale', 'closure')
        ON DUPLICATE KEY UPDATE reason=VALUES(reason)
    """, (closure_date,))

conn.commit()
print(f"✓ Inserite {len(company_closures)} chiusure aziendali")

# Calcola e inserisci saldi
for user_data in employees.values():
    for leave_type_id, leave_type_name in [(v, k) for k, v in leave_types.items()]:
        # Calcola giorni usati per questo tipo
        cursor.execute("""
            SELECT COALESCE(SUM(days), 0) 
            FROM leave_requests 
            WHERE userId = %s AND leaveTypeId = %s AND YEAR(STR_TO_DATE(startDate, '%%Y-%%m-%%d')) = 2025
        """, (user_data['id'], leave_type_id))
        used_days = cursor.fetchone()[0]
        
        # Inserisci saldo
        cursor.execute("""
            INSERT INTO leave_balances (user_id, leave_type_id, total_days, used_days, year)
            VALUES (%s, %s, 24, %s, 2025)
            ON DUPLICATE KEY UPDATE used_days = VALUES(used_days)
        """, (user_data['id'], leave_type_id, used_days))

conn.commit()
print(f"✓ Calcolati saldi ferie per {len(employees)} dipendenti")

cursor.close()
conn.close()

print("\n🎉 Importazione completata con successo!")
print(f"   - {len(employees)} dipendenti")
print(f"   - {total_requests} richieste di ferie")
print(f"   - {len(company_closures)} chiusure aziendali")
