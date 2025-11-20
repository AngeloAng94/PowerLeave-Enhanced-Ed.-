"""
Import selettivo: 10 dipendenti con ferie da marzo 2025 a gennaio 2026
"""
import openpyxl
import mysql.connector
from datetime import datetime
from collections import defaultdict
import os
from dotenv import load_dotenv

# Carica .env.import
load_dotenv('/home/ubuntu/dashboard-ferie-team/.env.import')

# Connessione database
conn = mysql.connector.connect(
    host=os.getenv('DB_HOST'),
    port=int(os.getenv('DB_PORT', 4000)),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    database=os.getenv('DB_NAME'),
    ssl_ca=os.getenv('DB_SSL_CA') if os.getenv('DB_SSL_CA') else None
)
cursor = conn.cursor()

# Apri Excel
wb = openpyxl.load_workbook('/home/ubuntu/dashboard-ferie-team/PianoFerie2025.xlsx')
ws = wb.active

# Trova date (riga 8, da colonna 3)
dates = {}
for col_idx in range(3, ws.max_column + 1):
    cell_value = ws.cell(8, col_idx).value
    if cell_value and isinstance(cell_value, datetime):
        date_str = cell_value.strftime('%Y-%m-%d')
        # Filtra solo marzo 2025 → gennaio 2026
        if '2025-03' <= date_str <= '2026-01-31':
            dates[col_idx] = date_str

print(f"✓ Trovate {len(dates)} date nel range marzo 2025 - gennaio 2026")

# Seleziona 10 dipendenti (righe 10-19)
employee_rows = list(range(10, 20))
employees = {}

for row_idx in employee_rows:
    name = ws.cell(row_idx, 2).value
    if name:
        # Crea utente
        cursor.execute(
            "INSERT INTO users (openId, name, email, role) VALUES (%s, %s, %s, %s) ON DUPLICATE KEY UPDATE name=VALUES(name)",
            (f"emp_{name.lower().replace(' ', '_')}", name, f"{name.lower().replace(' ', '.')}@company.com", "user")
        )
        conn.commit()
        
        # Get user ID
        cursor.execute("SELECT id FROM users WHERE openId = %s", (f"emp_{name.lower().replace(' ', '_')}",))
        user_id = cursor.fetchone()[0]
        employees[row_idx] = {'id': user_id, 'name': name}

print(f"✓ Creati {len(employees)} dipendenti")

# Get leave types
cursor.execute("SELECT id, name FROM leave_types")
leave_types = {name: id for id, name in cursor.fetchall()}
print(f"✓ Trovati {len(leave_types)} tipi di assenza")

# Raccogli ferie per dipendente
employee_leaves = defaultdict(lambda: defaultdict(list))

for row_idx, emp_data in employees.items():
    for date_col, date_str in dates.items():
        cell_value = ws.cell(row_idx, date_col).value
        if cell_value in ['X', 'P2', 'P4']:
            # Determina tipo e ore
            if cell_value == 'X':
                leave_type = 'FERIE/ROL 8H'
                hours = 8
            elif cell_value == 'P2':
                leave_type = 'PERMESSO 2 ORE'
                hours = 2
            else:  # P4
                leave_type = 'PERMESSO 4 ORE'
                hours = 4
            
            employee_leaves[emp_data['id']][leave_type].append((date_str, hours))

# Inserisci richieste raggruppate
total_requests = 0
for user_id, types_data in employee_leaves.items():
    for leave_type_name, dates_hours in types_data.items():
        leave_type_id = leave_types.get(leave_type_name)
        if not leave_type_id:
            continue
        
        # Raggruppa date consecutive
        dates_hours.sort()
        groups = []
        current_group = [dates_hours[0]]
        
        for i in range(1, len(dates_hours)):
            prev_date = datetime.strptime(current_group[-1][0], '%Y-%m-%d')
            curr_date = datetime.strptime(dates_hours[i][0], '%Y-%m-%d')
            diff = (curr_date - prev_date).days
            
            if diff == 1 and dates_hours[i][1] == current_group[-1][1]:
                current_group.append(dates_hours[i])
            else:
                groups.append(current_group)
                current_group = [dates_hours[i]]
        groups.append(current_group)
        
        # Inserisci ogni gruppo come richiesta
        for group in groups:
            start_date = group[0][0]
            end_date = group[-1][0]
            hours = group[0][1]
            days = len(group)
            
            cursor.execute("""
                INSERT INTO leave_requests 
                (userId, leaveTypeId, startDate, endDate, days, hours, status, notes)
                VALUES (%s, %s, %s, %s, %s, %s, 'approved', 'Importato da Excel')
            """, (user_id, leave_type_id, start_date, end_date, days, hours))
            total_requests += 1

conn.commit()
print(f"✓ Inserite {total_requests} richieste di ferie")

# Calcola e inserisci saldi
for user_id in employees.values():
    for leave_type_id, leave_type_name in [(v, k) for k, v in leave_types.items()]:
        # Calcola giorni usati per questo tipo
        cursor.execute("""
            SELECT COALESCE(SUM(days), 0) 
            FROM leave_requests 
            WHERE userId = %s AND leaveTypeId = %s AND YEAR(STR_TO_DATE(startDate, '%%Y-%%m-%%d')) = 2025
        """, (user_id['id'], leave_type_id))
        used_days = cursor.fetchone()[0]
        
        # Inserisci saldo
        cursor.execute("""
            INSERT INTO leave_balances (user_id, leave_type_id, total_days, used_days, year)
            VALUES (%s, %s, 24, %s, 2025)
            ON DUPLICATE KEY UPDATE used_days = VALUES(used_days)
        """, (user_id['id'], leave_type_id, used_days))

conn.commit()
print(f"✓ Calcolati saldi ferie per {len(employees)} dipendenti")

cursor.close()
conn.close()

print("\n🎉 Importazione completata con successo!")
