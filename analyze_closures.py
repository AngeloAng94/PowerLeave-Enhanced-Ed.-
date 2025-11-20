import openpyxl
from collections import defaultdict
from datetime import datetime

wb = openpyxl.load_workbook('/home/ubuntu/dashboard-ferie-team/PianoFerie2025.xlsx')
ws = wb.active

# Trova intestazione date (riga 8, da colonna 3 in poi)
date_row = 8
dates = {}
for col_idx in range(3, ws.max_column + 1):
    cell_value = ws.cell(date_row, col_idx).value
    if cell_value and isinstance(cell_value, datetime):
        dates[col_idx] = cell_value.strftime('%Y-%m-%d')

print(f"✓ Trovate {len(dates)} date")

# Trova dipendenti (righe 10-39, 1 riga per dipendente)
employee_rows = list(range(10, 40))
total_employees = len(employee_rows)
print(f"✓ Trovati {total_employees} dipendenti")

# Conta "X", "P2", "P4" per ogni data
date_counts = defaultdict(int)
for date_col in dates.keys():
    for emp_row in employee_rows:
        cell_value = ws.cell(emp_row, date_col).value
        if cell_value in ['X', 'P2', 'P4']:
            date_counts[dates[date_col]] += 1

# Identifica chiusure (>= 80% dipendenti)
threshold = total_employees * 0.8
closures = []
for date, count in sorted(date_counts.items()):
    if count >= threshold:
        percentage = (count / total_employees) * 100
        closures.append((date, count, percentage))

print(f"\n🏢 Trovate {len(closures)} chiusure aziendali (>= 80% dipendenti):\n")
for date, count, pct in closures[:20]:  # Mostra prime 20
    print(f"  {date}: {count}/{total_employees} dipendenti ({pct:.0f}%)")
if len(closures) > 20:
    print(f"\n  ... e altre {len(closures) - 20} chiusure")
