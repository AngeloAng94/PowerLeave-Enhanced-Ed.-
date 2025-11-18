#!/usr/bin/env python3
"""
Script di importazione dati da Piano Ferie Excel al database della dashboard.

Questo script:
1. Legge il file Excel PianoFerie2025.xlsx
2. Estrae i nomi dei dipendenti
3. Identifica le date con "X" (ferie/permessi)
4. Crea utenti nel database (se non esistono)
5. Crea richieste di ferie con stato "approved"
6. Calcola e aggiorna i saldi ferie

Uso:
    python3 import_excel_to_db.py /path/to/PianoFerie2025.xlsx

Requisiti:
    pip install openpyxl mysql-connector-python python-dotenv
"""

import openpyxl
import mysql.connector
from datetime import datetime
from collections import defaultdict
import sys
import os
from dotenv import load_dotenv

# Carica variabili d'ambiente
load_dotenv()

# Configurazione database
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'dashboard_ferie'),
}

# Mapping tipi di assenza (riga Excel -> nome tipo)
LEAVE_TYPE_MAPPING = {
    3: "FERIE/ROL 8 H",
    4: "PERMESSO 2 ORE",
    5: "PERMESSO 4 ORE",
}

# Mapping ore per tipo
HOURS_MAPPING = {
    "FERIE/ROL 8 H": 8,
    "PERMESSO 2 ORE": 2,
    "PERMESSO 4 ORE": 4,
}


def connect_db():
    """Connessione al database MySQL"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        print(f"✓ Connesso al database: {DB_CONFIG['database']}")
        return conn
    except mysql.connector.Error as err:
        print(f"✗ Errore connessione database: {err}")
        sys.exit(1)


def get_or_create_leave_type(cursor, name, hours):
    """Ottiene o crea un tipo di assenza"""
    cursor.execute("SELECT id FROM leave_types WHERE name = %s", (name,))
    result = cursor.fetchone()
    
    if result:
        return result[0]
    
    # Crea nuovo tipo
    cursor.execute(
        "INSERT INTO leave_types (name, description, requiresApproval) VALUES (%s, %s, %s)",
        (name, f"Importato da Excel - {hours} ore", True)
    )
    return cursor.lastrowid


def get_or_create_user(cursor, name):
    """Ottiene o crea un utente"""
    # Cerca per nome
    cursor.execute("SELECT id FROM users WHERE name = %s", (name,))
    result = cursor.fetchone()
    
    if result:
        return result[0]
    
    # Crea nuovo utente
    # Genera openId univoco basato sul nome
    open_id = f"excel_import_{name.lower().replace(' ', '_')}"
    email = f"{name.lower().replace(' ', '.')}@example.com"
    
    cursor.execute(
        """INSERT INTO users (openId, name, email, loginMethod, role) 
           VALUES (%s, %s, %s, %s, %s)""",
        (open_id, name, email, "excel_import", "user")
    )
    return cursor.lastrowid


def parse_excel(file_path):
    """Analizza il file Excel e estrae i dati delle ferie"""
    print(f"\n📂 Apertura file: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Piano Ferie']
    
    print(f"✓ Foglio caricato: {ws.max_row} righe × {ws.max_column} colonne")
    
    # Estrai dipendenti (colonna 2, righe 10+)
    employees = {}
    for row_idx in range(10, ws.max_row + 1):
        name = ws.cell(row_idx, 2).value
        if name and isinstance(name, str) and name.strip():
            employees[row_idx] = name.strip()
    
    print(f"✓ Trovati {len(employees)} dipendenti")
    for row_idx, name in employees.items():
        print(f"  - Riga {row_idx}: {name}")
    
    # Estrai date (riga 9, colonna 3+)
    dates = {}
    for col_idx in range(3, ws.max_column + 1):
        date_val = ws.cell(9, col_idx).value
        if isinstance(date_val, datetime):
            dates[col_idx] = date_val.strftime("%Y-%m-%d")
    
    print(f"✓ Trovate {len(dates)} date (da {min(dates.values())} a {max(dates.values())})")
    
    # Estrai le "X" (ferie/permessi)
    leave_records = []
    
    for emp_row, emp_name in employees.items():
        employee_leaves = defaultdict(list)  # data -> lista di tipi
        
        for col_idx, date_str in dates.items():
            cell_value = ws.cell(emp_row, col_idx).value
            
            if cell_value and str(cell_value).strip().upper() == "X":
                # Determina il tipo di assenza dalla riga delle legende (3-5)
                # Per ora assumiamo che tutte le X siano "FERIE/ROL 8 H"
                # In futuro si potrebbe usare la formattazione o altri indicatori
                leave_type = "FERIE/ROL 8 H"
                hours = 8
                
                employee_leaves[date_str].append({
                    'type': leave_type,
                    'hours': hours
                })
        
        # Converti in richieste (raggruppa date consecutive)
        for date_str, types in employee_leaves.items():
            for leave_info in types:
                leave_records.append({
                    'employee': emp_name,
                    'date': date_str,
                    'leave_type': leave_info['type'],
                    'hours': leave_info['hours']
                })
    
    print(f"✓ Trovate {len(leave_records)} giornate di ferie/permessi")
    
    return employees, leave_records


def group_consecutive_dates(leave_records):
    """Raggruppa date consecutive per lo stesso dipendente e tipo"""
    # Ordina per dipendente, tipo, data
    sorted_records = sorted(
        leave_records,
        key=lambda x: (x['employee'], x['leave_type'], x['date'])
    )
    
    grouped = []
    current_group = None
    
    for record in sorted_records:
        if current_group is None:
            current_group = {
                'employee': record['employee'],
                'leave_type': record['leave_type'],
                'hours': record['hours'],
                'start_date': record['date'],
                'end_date': record['date'],
                'days': 1
            }
        elif (current_group['employee'] == record['employee'] and
              current_group['leave_type'] == record['leave_type'] and
              current_group['hours'] == record['hours']):
            # Stessa persona, stesso tipo, stesse ore -> estendi periodo
            current_group['end_date'] = record['date']
            current_group['days'] += 1
        else:
            # Nuovo gruppo
            grouped.append(current_group)
            current_group = {
                'employee': record['employee'],
                'leave_type': record['leave_type'],
                'hours': record['hours'],
                'start_date': record['date'],
                'end_date': record['date'],
                'days': 1
            }
    
    if current_group:
        grouped.append(current_group)
    
    return grouped


def import_to_database(employees, leave_records, dry_run=False):
    """Importa i dati nel database"""
    conn = connect_db()
    cursor = conn.cursor()
    
    try:
        print("\n📥 Inizio importazione...")
        
        # 1. Crea/ottieni tipi di assenza
        print("\n1️⃣  Gestione tipi di assenza...")
        leave_type_ids = {}
        for type_name, hours in HOURS_MAPPING.items():
            type_id = get_or_create_leave_type(cursor, type_name, hours)
            leave_type_ids[type_name] = type_id
            print(f"  ✓ {type_name} (ID: {type_id})")
        
        # 2. Crea/ottieni utenti
        print("\n2️⃣  Gestione utenti...")
        user_ids = {}
        for emp_name in set(emp['employee'] for emp in leave_records):
            user_id = get_or_create_user(cursor, emp_name)
            user_ids[emp_name] = user_id
            print(f"  ✓ {emp_name} (ID: {user_id})")
        
        # 3. Raggruppa date consecutive
        print("\n3️⃣  Raggruppamento richieste consecutive...")
        grouped_requests = group_consecutive_dates(leave_records)
        print(f"  ✓ {len(leave_records)} giorni → {len(grouped_requests)} richieste")
        
        # 4. Inserisci richieste
        print("\n4️⃣  Inserimento richieste nel database...")
        inserted = 0
        
        for request in grouped_requests:
            user_id = user_ids[request['employee']]
            type_id = leave_type_ids[request['leave_type']]
            
            if not dry_run:
                cursor.execute(
                    """INSERT INTO leave_requests 
                       (userId, leaveTypeId, startDate, endDate, days, hours, status, createdAt, updatedAt)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())""",
                    (user_id, type_id, request['start_date'], request['end_date'], 
                     request['days'], request['hours'], 'approved')
                )
                inserted += 1
                
                if inserted % 50 == 0:
                    print(f"  ... {inserted} richieste inserite")
        
        print(f"  ✓ {inserted} richieste inserite con successo")
        
        # 5. Calcola saldi ferie
        print("\n5️⃣  Calcolo saldi ferie...")
        current_year = datetime.now().year
        
        for emp_name, user_id in user_ids.items():
            # Calcola giorni utilizzati per questo utente
            cursor.execute(
                """SELECT SUM(days) FROM leave_requests 
                   WHERE userId = %s AND status = 'approved' 
                   AND YEAR(STR_TO_DATE(startDate, '%Y-%m-%d')) = %s""",
                (user_id, current_year)
            )
            result = cursor.fetchone()
            used_days = result[0] if result[0] else 0
            
            # Assegna saldo standard (es. 22 giorni all'anno)
            total_days = 22
            available_days = total_days - used_days
            
            if not dry_run:
                # Verifica se esiste già un saldo
                cursor.execute(
                    "SELECT id FROM leave_balances WHERE userId = %s AND year = %s",
                    (user_id, current_year)
                )
                existing = cursor.fetchone()
                
                if existing:
                    cursor.execute(
                        """UPDATE leave_balances 
                           SET totalDays = %s, usedDays = %s, availableDays = %s
                           WHERE id = %s""",
                        (total_days, used_days, available_days, existing[0])
                    )
                else:
                    cursor.execute(
                        """INSERT INTO leave_balances 
                           (userId, year, totalDays, usedDays, availableDays)
                           VALUES (%s, %s, %s, %s, %s)""",
                        (user_id, current_year, total_days, used_days, available_days)
                    )
            
            print(f"  ✓ {emp_name}: {used_days}/{total_days} giorni utilizzati")
        
        if not dry_run:
            conn.commit()
            print("\n✅ Importazione completata con successo!")
        else:
            print("\n⚠️  DRY RUN - Nessuna modifica effettuata al database")
        
    except Exception as e:
        conn.rollback()
        print(f"\n❌ Errore durante l'importazione: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def main():
    """Funzione principale"""
    if len(sys.argv) < 2:
        print("Uso: python3 import_excel_to_db.py <file_excel.xlsx> [--dry-run]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    dry_run = "--dry-run" in sys.argv
    
    if not os.path.exists(excel_file):
        print(f"❌ File non trovato: {excel_file}")
        sys.exit(1)
    
    print("=" * 80)
    print("IMPORTAZIONE PIANO FERIE DA EXCEL")
    print("=" * 80)
    
    if dry_run:
        print("\n⚠️  MODALITÀ DRY RUN - Nessuna modifica verrà effettuata")
    
    # Analizza Excel
    employees, leave_records = parse_excel(excel_file)
    
    # Importa nel database
    import_to_database(employees, leave_records, dry_run=dry_run)
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
